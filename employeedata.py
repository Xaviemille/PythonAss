#This statement is used for importing a csv module that is used for parsing tabular like data structure
import csv

#This provides functions for creating and removing a directory (folder)
import os

#This function is used when you have to access an MS Excel file in openpyxl module.
from openpyxl import workbook, load_workbook 

if os.path.exists("employeedata.csv" and "employeedata.xlsx"):
   #This function is used to open the file employeedata.csv
    csv_file = open("employeedata.csv", "r")
#"helpinghands.cm" is replaced by "handsinhands.org" and is returned where "helpinghands.cm" have been joined by str separator.
    csv_file = ''.join([i for i in csv_file]).replace("helpinghands.cm", "handsinhands.org")
#A new file "newemail" is created in write mode
    newcsv_file = open("newemployeedata.csv","w")
#The content of the file "employeedata.csv" is written and the modifications added
    newcsv_file.writelines(csv_file)
#The file is closed
    newcsv_file.close()
    #Here, we are specifying we want to access the database of employee we created in excel by passing it's name
    wb = load_workbook('employeedata.xlsx')
    #A worksheet is created
    ws = wb.active
    #Here, we are giving a count of the number of occupied rows and stating we want to start from the second row
    for i in range(2,ws.max_row+1):
    #This is to remove the previous data that has been saved
        cell= ws.cell(i, 2)
    #Here, anywhere 'helpinghands.cm' will be found in the rows of the excel document, it will be replaced by 'handsinhands'
        if 'helpinghands.cm' in cell.value:
            #Anywhere 'helpinghands.cm' will be found in the file, it will be replaced by 'handsinhands.org'
            newemployeedata=(cell.value).replace('helpinghands.cm', 'handsinhands.org')
            ws.cell(i,2).value=newemployeedata
    #A new excel document is created with modifications added
    wb.save('newemployeedata.xlsx')


    

#If the file employeedata.csv exists, execute the code
if os.path.exists("./employeedata.csv") == True:
#This function is used to open the file employeedata.csv
    csv_file = open("employeedata.csv", "r")
#"helpinghands.cm" is replaced by "handsinhands.org" and is returned where "helpinghands.cm" have been joined by str separator.
    csv_file = ''.join([i for i in csv_file]).replace("helpinghands.cm", "handsinhands.org")
#A new file "newemail" is created in write mode
    newcsv_file = open("newemployeedata.csv","w")
#The content of the file "employeedata.csv" is written and the modifications added
    newcsv_file.writelines(csv_file)
#The file is closed
    newcsv_file.close()

#Also, if the file employeedata.xlsx exists, execute the code
elif os.path.exists("./employeedata.xlsx") == True:
    #Here, we are specifying we want to access the database of employee we created in excel by passing it's name
    wb = load_workbook('employeedata.xlsx')
    #A worksheet is created
    ws = wb.active
    #Here, we are giving a count of the number of occupied rows and stating we want to start from the second row
    for i in range(2,ws.max_row+1):
    #This is to remove the previous data that has been saved
        cell= ws.cell(i, 2)
    #Here, anywhere 'helpinghands.cm' will be found in the rows of the excel document, it will be replaced by 'handsinhands'
        if 'helpinghands.cm' in cell.value:
            #Anywhere 'helpinghands.cm' will be found in the file, it will be replaced by 'handsinhands.org'
            newemployeedata=(cell.value).replace('helpinghands.cm', 'handsinhands.org')
            ws.cell(i,2).value=newemployeedata
    #A new excel document is created with modifications added
    wb.save('newemployeedata.xlsx')

#In case either "employeedata.csv" or "employeedata.xlsx" is not found, a message will be displayed
else:
    print ("No file exist")